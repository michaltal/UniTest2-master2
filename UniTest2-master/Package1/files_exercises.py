from openpyxl import Workbook
from openpyxl import load_workbook


# Exercise 14.1 a
with open('story.txt', 'r+') as file:
    L1 = 'A boy is playing there.\n'
    L2 = 'There is a playground.\n'
    L3 = 'An airplane is in the sky.\n'
    L4 = 'The sky is pink.\n'
    L5 = 'Alphabets and numbers are allowed in the password.\n'
    L = [L1, L2, L3, L4, L5]
    file.writelines(L)

# Exercise 14.1 b
with open('story.txt', 'r') as file:
    print("The file is:\n", file.read())


# Exercise 14.2 a
def count_lines_first_letter():
    file1 = open('story.txt', 'r')
    count = 0
    for line in file1:
        if line[0] not in 'T':
            count += 1
    file1.close()
    print("No of lines NOT starting with 'T'=", count)


count_lines_first_letter()


# Exercise 14.2 b
def count_words_in_lines():
    file2 = open('story.txt', 'r')
    count = 0
    for line in file2:
        words = line.split()
        # print(words)
        for word in words:
            if word == "the" or word == "The":
                count += 1
                # print(word)
    print("No of times the word THE appears in the file = ",count)


count_words_in_lines()


# Exercise 14.2 c
def count_words_in_file():
    file3 = open('story.txt', 'r')
    count = 0
    for line in file3:
        words = line.split()
        for word in words:
            count += 1
    print("No of words in the file = ",count)


count_words_in_file()


# Exercise 14.3 a
def read_line_by_line():
    print('read lines:')
    with open('story.txt', 'r') as file:
        # print(file.read())
        count = 0
        while True:
            count += 1
            line = file.readline()

            # if line is empty
            # end of file is reached
            if not line:
                break
            print("Line num {}: {}".format(count, line.strip()))


read_line_by_line()


# Exercise 14.3 b
def num_of_lines():
    with open('story.txt', 'r') as file:
        count = 0
        while True:
            count += 1
            line = file.readline()

            # if line is empty
            # end of file is reached
            if not line:
                break
        print("number of Lines in the file: ",count-1)


num_of_lines()

# Exercise 14.4 a
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "UserName"
sheet["B1"] = "Password"
sheet["A2"] = "Tester1"
sheet["B2"] = "1234"
sheet["A3"] = "Tester2"
sheet["B3"] = "45678"
sheet["F10"] = "Breath"

workbook.save("Login.xlsx")

# Exercise 14.4 b
workbook = load_workbook("Login.xlsx")
# print("Excel Login Sheets Names: ",workbook.sheetnames)
sheet = workbook.active
# print(sheet)
print(sheet.title)
print(sheet["A2"].value)
print(sheet["B2"].value)
print(sheet.cell(row=10, column=6).value)
sheet["A1:B2"]

# Exercise 14.4 c
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2, values_only=True):
    print("Rows: ",row)


for column in sheet.iter_cols(min_row=1, max_row=3, min_col=1, max_col=2, values_only=True):
    print("Columns: ",column)









