from openpyxl import load_workbook

workbook = load_workbook(filename="storedata.xlsx")
sheet = workbook.active
sum_all = 0
average = 0
count = 0

for row in sheet.iter_rows(min_row=2, max_row=7, min_col=4, max_col=4, values_only=True):
    for cell in row:
        print("cell: ", cell)
        sum_all = sum_all + cell
        count +=1
average = sum_all/count
print("Sum :", sum_all)
print("average :", average)
print("count :", count)


